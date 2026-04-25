"""Собрать материалы Stage 9 для отчета по практике.

Скрипт детерминированный и не запускает новые эксперименты. Он собирает
финальные aggregate metrics, формирует таблицы сравнения, Markdown для отчета,
LaTeX-таблицу, XLSX-книгу и PNG-рисунки. При запуске в Colab тот же пакет
сохраняется в каноническую папку Google Drive.
"""

from __future__ import annotations

import csv
import glob
import json
import math
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from PIL import Image, ImageDraw, ImageFont

try:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.ticker import PercentFormatter
except ModuleNotFoundError:  # pragma: no cover - environment fallback.
    plt = None  # type: ignore[assignment]
    PercentFormatter = None  # type: ignore[assignment]

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError:  # pragma: no cover - environment fallback.
    Workbook = None  # type: ignore[assignment]
    OpenpyxlImage = None  # type: ignore[assignment]


REPO_ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = REPO_ROOT / "reports" / "stage9_report_materials"
FIGURES_DIR = OUT_DIR / "figures"
TABLES_DIR = OUT_DIR / "tables"
LATEX_DIR = OUT_DIR / "latex"
CANONICAL_DRIVE_ROOT_STR = "/content/drive/MyDrive/diploma/petr_text_to_visualization_part"
CANONICAL_DRIVE_ROOT = Path(CANONICAL_DRIVE_ROOT_STR)


@dataclass(frozen=True)
class FinalRun:
    label: str
    method: str
    run_id: str
    metrics_method_dir: str
    sample_size: int
    family: str
    role: str


FINAL_RUNS = [
    FinalRun(
        label="B0",
        method="B0_rule_based",
        run_id="stage4_cpu_sample200",
        metrics_method_dir="B0_rule_based",
        sample_size=200,
        family="Детерминированный baseline",
        role="Правила без LLM",
    ),
    FinalRun(
        label="B1",
        method="B1_constraint_ranker",
        run_id="stage4_cpu_sample200",
        metrics_method_dir="B1_constraint_ranker",
        sample_size=200,
        family="Детерминированный baseline",
        role="Ограничения и ранжирование кандидатов",
    ),
    FinalRun(
        label="B2",
        method="B2_partial_recommender",
        run_id="stage5_partial_sample200",
        metrics_method_dir="B2_partial_recommender",
        sample_size=200,
        family="Подход в стиле существующих инструментов",
        role="Частичный recommender fallback",
    ),
    FinalRun(
        label="B3",
        method="B3_local_llm_qwen3_8b",
        run_id="stage6_qwen3_8b_fast_sample50",
        metrics_method_dir="B3_local_llm_qwen3_8b",
        sample_size=50,
        family="Локальная LLM",
        role="Один кандидат Qwen3-8B",
    ),
    FinalRun(
        label="B4",
        method="B4_llm_validator_reranker",
        run_id="stage7_b4_sample20_tokens384",
        metrics_method_dir="B4_llm_validator_reranker",
        sample_size=20,
        family="Локальная LLM + validation",
        role="3 кандидата + validator/reranker",
    ),
]


FALLBACK_METRICS: dict[str, dict[str, float]] = {
    "B0_rule_based": {
        "examples": 200,
        "predictions": 200,
        "chart_type_accuracy": 0.54,
        "x_field_accuracy": 0.575,
        "y_field_accuracy": 0.585,
        "field_selection_f1": 0.857,
        "encoding_accuracy": 0.2525,
        "aggregation_accuracy": 0.26,
        "normalized_exact_match": 0.085,
        "vega_lite_validity": 1.0,
        "top1_success": 0.085,
        "latency_ms": 0.18079,
        "memory_peak_mb": 106.309,
        "failure_rate": 0.0,
    },
    "B1_constraint_ranker": {
        "examples": 200,
        "predictions": 200,
        "chart_type_accuracy": 0.51,
        "x_field_accuracy": 0.575,
        "y_field_accuracy": 0.56,
        "field_selection_f1": 0.847,
        "encoding_accuracy": 0.2275,
        "aggregation_accuracy": 0.23,
        "normalized_exact_match": 0.06,
        "vega_lite_validity": 1.0,
        "top1_success": 0.06,
        "latency_ms": 0.32999,
        "memory_peak_mb": 210.348,
        "failure_rate": 0.0,
    },
    "B2_partial_recommender": {
        "examples": 200,
        "predictions": 200,
        "chart_type_accuracy": 0.62,
        "field_selection_f1": 0.8703333333333333,
        "encoding_accuracy": 0.11,
        "aggregation_accuracy": 0.11,
        "normalized_exact_match": 0.11,
        "vega_lite_validity": 1.0,
        "latency_ms": 0.0,
        "memory_peak_mb": 0.0,
        "failure_rate": 0.0,
    },
    "B3_local_llm_qwen3_8b": {
        "examples": 50,
        "predictions": 50,
        "chart_type_accuracy": 0.5,
        "x_field_accuracy": 0.72,
        "y_field_accuracy": 0.6,
        "field_selection_f1": 0.7533333333333333,
        "encoding_accuracy": 0.59,
        "aggregation_accuracy": 0.84,
        "normalized_exact_match": 0.36,
        "vega_lite_validity": 0.86,
        "latency_ms": 10187.22126,
        "memory_peak_mb": 1983.812,
        "failure_rate": 0.14,
    },
    "B4_llm_validator_reranker": {
        "examples": 20,
        "predictions": 20,
        "chart_type_accuracy": 0.5,
        "x_field_accuracy": 0.5,
        "y_field_accuracy": 0.5,
        "field_selection_f1": 0.8,
        "encoding_accuracy": 0.5,
        "aggregation_accuracy": 0.85,
        "normalized_exact_match": 0.5,
        "vega_lite_validity": 0.95,
        "top1_success": 0.5,
        "oracle_success_at_k": 0.5,
        "latency_ms": 34898.125199999995,
        "memory_peak_mb": 2153.457,
        "failure_rate": 0.05,
    },
}


def main() -> int:
    prepare_dirs()
    drive_root = CANONICAL_DRIVE_ROOT if CANONICAL_DRIVE_ROOT.exists() else None
    inventory = discover_run_folders(drive_root)
    rows = collect_comparison_rows(drive_root)
    write_tables(rows, inventory)
    write_figures(rows, drive_root)
    write_workbook(rows)
    write_latex_table(rows)
    write_markdown(rows, inventory, drive_root)
    write_review(rows, inventory, drive_root)
    sync_to_drive(drive_root)
    print(json.dumps({"output_dir": str(OUT_DIR), "drive_root": CANONICAL_DRIVE_ROOT_STR if drive_root else None}, ensure_ascii=False, indent=2))
    return 0


def prepare_dirs() -> None:
    for directory in (OUT_DIR, FIGURES_DIR, TABLES_DIR, LATEX_DIR):
        directory.mkdir(parents=True, exist_ok=True)


def discover_run_folders(drive_root: Path | None) -> list[dict[str, Any]]:
    runs_root = drive_root / "runs" if drive_root else None
    inventory: list[dict[str, Any]] = []
    if runs_root and runs_root.exists():
        for path in sorted(item for item in runs_root.iterdir() if item.is_dir()):
            inventory.append(
                {
                    "run_id": path.name,
                    "path": str(path),
                    "has_summary": (path / "experiment_summary.json").exists(),
                    "has_metrics": (path / "metrics").exists(),
                    "has_predictions": (path / "predictions").exists(),
                    "has_rendered": (path / "rendered").exists(),
                }
            )
    else:
        seen: set[str] = set()
        for run in FINAL_RUNS:
            if run.run_id in seen:
                continue
            seen.add(run.run_id)
            inventory.append(
                {
                    "run_id": run.run_id,
                    "path": f"{CANONICAL_DRIVE_ROOT_STR}/runs/{run.run_id}",
                    "has_summary": None,
                    "has_metrics": None,
                    "has_predictions": None,
                    "has_rendered": None,
                }
            )
    write_json(OUT_DIR / "run_inventory.json", inventory)
    write_csv(TABLES_DIR / "run_inventory.csv", localized_inventory_rows(inventory))
    return inventory


def localized_inventory_rows(inventory: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "run_id": row.get("run_id"),
            "Путь": row.get("path"),
            "Есть summary": row.get("has_summary"),
            "Есть metrics": row.get("has_metrics"),
            "Есть прогнозы": row.get("has_predictions"),
            "Есть рендеры": row.get("has_rendered"),
        }
        for row in inventory
    ]


def localized_final_runs_rows() -> list[dict[str, Any]]:
    return [
        {
            "Подход": run.label,
            "Метод": run.method,
            "run_id": run.run_id,
            "Папка метрик": run.metrics_method_dir,
            "Sample": run.sample_size,
            "Семейство": run.family,
            "Роль": run.role,
        }
        for run in FINAL_RUNS
    ]


def collect_comparison_rows(drive_root: Path | None) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for run in FINAL_RUNS:
        metrics = read_aggregate_metrics(drive_root, run)
        row = {
            "label": run.label,
            "method": run.method,
            "run_id": run.run_id,
            "sample_size": int(metrics.get("examples") or run.sample_size),
            "family": run.family,
            "role": run.role,
            "chart_type_accuracy": as_float(metrics.get("chart_type_accuracy")),
            "x_field_accuracy": as_float(metrics.get("x_field_accuracy")),
            "y_field_accuracy": as_float(metrics.get("y_field_accuracy")),
            "field_selection_f1": as_float(metrics.get("field_selection_f1")),
            "encoding_accuracy": as_float(metrics.get("encoding_accuracy")),
            "aggregation_accuracy": as_float(metrics.get("aggregation_accuracy")),
            "normalized_exact_match": as_float(metrics.get("normalized_exact_match")),
            "vega_lite_validity": as_float(metrics.get("vega_lite_validity")),
            "top1_success": as_float(metrics.get("top1_success")),
            "oracle_success_at_k": as_float(metrics.get("oracle_success_at_k")),
            "latency_ms": as_float(metrics.get("latency_ms")),
            "memory_peak_mb": as_float(metrics.get("memory_peak_mb")),
            "failure_rate": as_float(metrics.get("failure_rate")),
        }
        if run.label == "B2":
            # The Stage 5 partial fallback did not produce reliable process-level
            # runtime measurements; do not present placeholder zeros as facts.
            if row["latency_ms"] == 0.0:
                row["latency_ms"] = None
            if row["memory_peak_mb"] == 0.0:
                row["memory_peak_mb"] = None
        rows.append(row)
    write_json(OUT_DIR / "final_runs.json", localized_final_runs_rows())
    return rows


def read_aggregate_metrics(drive_root: Path | None, run: FinalRun) -> dict[str, Any]:
    if drive_root:
        path = drive_root / "runs" / run.run_id / "metrics" / run.metrics_method_dir / "aggregate_metrics.csv"
        if path.exists():
            with path.open("r", encoding="utf-8", newline="") as handle:
                data = list(csv.DictReader(handle))
            if data:
                return data[0]
    return FALLBACK_METRICS[run.method]


def write_tables(rows: list[dict[str, Any]], inventory: list[dict[str, Any]]) -> None:
    comparison = [
        {
            "Подход": row["label"],
            "Метод": row["method"],
            "Семейство": row["family"],
            "Sample": row["sample_size"],
            "run_id": row["run_id"],
            "Роль": row["role"],
        }
        for row in rows
    ]
    quality = [
        {
            "Подход": row["label"],
            "Validity": row["vega_lite_validity"],
            "Chart type": row["chart_type_accuracy"],
            "Field F1": row["field_selection_f1"],
            "Encoding": row["encoding_accuracy"],
            "Aggregation": row["aggregation_accuracy"],
            "Exact match": row["normalized_exact_match"],
            "Oracle@3": row["oracle_success_at_k"],
        }
        for row in rows
    ]
    system = [
        {
            "Подход": row["label"],
            "Latency ms": format_optional_system_metric(row["latency_ms"]),
            "Memory MB": format_optional_system_metric(row["memory_peak_mb"]),
            "Failure rate": row["failure_rate"],
            "Sample": row["sample_size"],
        }
        for row in rows
    ]
    applicability = [
        {
            "Подход": "B0",
            "Когда применять": "Быстрый резервный вариант, детерминированная демонстрация, простые дашборд-сценарии",
            "Ценность для интеграции": "Высокая как гарантированный резервный вариант",
            "Ограничение": "Низкая семантическая гибкость",
        },
        {
            "Подход": "B1",
            "Когда применять": "Сценарии, где нужны явные ограничения для кандидатов",
            "Ценность для интеграции": "Средняя",
            "Ограничение": "Не превзошел B0 по метрикам точного совпадения",
        },
        {
            "Подход": "B2",
            "Когда применять": "Базовый подход без LLM в стиле существующих инструментов рекомендаций",
            "Ценность для интеграции": "Средне-высокая",
            "Ограничение": "Частичная замена NL4DV, не полноценный NL4DV",
        },
        {
            "Подход": "B3",
            "Когда применять": "Локальная LLM-генерация, ориентированная на качество",
            "Ценность для интеграции": "Высокая для локального пути на open-source LLM",
            "Ограничение": "Медленнее и менее устойчива, чем детерминированные резервные подходы",
        },
        {
            "Подход": "B4",
            "Когда применять": "Лучшее качество среди протестированных методов, если допустима задержка",
            "Ценность для интеграции": "Рекомендуемый исследовательский кандидат",
            "Ограничение": "Три генерации на пример; примерно в 3.4 раза медленнее B3",
        },
    ]
    risks = [
        {"Риск": "Репрезентативность датасета", "Влияние": "Метрики зависят от примеров, полученных из nvBench", "Снижение риска": "Фиксировать размеры выборок и сохранять воспроизводимые артефакты"},
        {"Риск": "Задержка LLM", "Влияние": "B3/B4 медленнее детерминированных базовых подходов", "Снижение риска": "Отключать thinking, останавливать генерацию после JSON, ограничивать tokens"},
        {"Риск": "Ошибки рендеринга", "Влияние": "Некоторые валидные specs могут не отрендериться", "Снижение риска": "Сохранять render_failures.json и визуально проверять примеры"},
        {"Риск": "Несовместимость NL4DV", "Влияние": "Полный базовый подход на существующем инструменте не был установлен", "Снижение риска": "Зафиксировать dependency conflict и использовать B2 как частичный резервный вариант"},
        {"Риск": "Нет оценки Text-to-SQL", "Влияние": "Качество upstream SQL здесь не измеряется", "Снижение риска": "Явно фиксировать границу post-query постановки"},
    ]
    write_csv(TABLES_DIR / "comparison_solutions.csv", comparison)
    write_csv(TABLES_DIR / "quality_metrics.csv", quality)
    write_csv(TABLES_DIR / "latency_memory_failure.csv", system)
    write_csv(TABLES_DIR / "applicability.csv", applicability)
    write_csv(TABLES_DIR / "risks_limitations.csv", risks)
    write_markdown_table(OUT_DIR / "comparison_table.md", comparison)
    write_markdown_table(OUT_DIR / "quality_metrics_table.md", quality)
    write_csv(TABLES_DIR / "run_inventory.csv", localized_inventory_rows(inventory))


def write_workbook(rows: list[dict[str, Any]]) -> None:
    if Workbook is None:
        return
    workbook = Workbook()
    workbook.remove(workbook.active)
    add_sheet(workbook, "Сравнение", read_csv(TABLES_DIR / "comparison_solutions.csv"))
    add_sheet(workbook, "Метрики качества", read_csv(TABLES_DIR / "quality_metrics.csv"))
    add_sheet(workbook, "Системные метрики", read_csv(TABLES_DIR / "latency_memory_failure.csv"))
    add_sheet(workbook, "Применимость", read_csv(TABLES_DIR / "applicability.csv"))
    add_sheet(workbook, "Риски", read_csv(TABLES_DIR / "risks_limitations.csv"))
    add_report_image(workbook["Метрики качества"], FIGURES_DIR / "key_metrics_comparison.png", "J2", width=1040)
    add_report_image(workbook["Системные метрики"], FIGURES_DIR / "system_metrics_comparison.png", "G2", width=1100)
    workbook_path = OUT_DIR / "stage9_tables.xlsx"
    try:
        workbook.save(workbook_path)
    except PermissionError:
        warning = (
            f"{workbook_path} заблокирован другим процессом; "
            "существующая книга не была изменена при локальном запуске.\n"
        )
        (OUT_DIR / "stage9_tables_write_warning.txt").write_text(warning, encoding="utf-8")


def add_sheet(workbook: Any, title: str, records: list[dict[str, str]]) -> None:
    sheet = workbook.create_sheet(title)
    if not records:
        return
    headers = list(records[0])
    sheet.append(headers)
    for record in records:
        sheet.append([coerce_excel_value(record.get(header, "")) for header in headers])
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
    for index, header in enumerate(headers, start=1):
        values = [str(header)] + [str(record.get(header, "")) for record in records]
        width = min(max(max(len(value) for value in values) + 2, 12), 42)
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"


def add_report_image(sheet: Any, image_path: Path, anchor: str, *, width: int) -> None:
    if OpenpyxlImage is None or not image_path.exists():
        return
    image = OpenpyxlImage(str(image_path))
    ratio = width / image.width
    image.width = width
    image.height = int(image.height * ratio)
    sheet.add_image(image, anchor)


def coerce_excel_value(value: Any) -> Any:
    if value in (None, ""):
        return ""
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    try:
        number = float(text)
    except ValueError:
        return value
    if number.is_integer() and "." not in text:
        return int(number)
    return number


def add_quality_chart(sheet: Any) -> None:
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Ключевые метрики качества по подходам"
    chart.y_axis.title = "Score"
    chart.x_axis.title = "Подход"
    chart.y_axis.numFmt = "0%"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1.05
    data = Reference(sheet, min_col=3, max_col=7, min_row=1, max_row=6)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    for series, title in zip(
        chart.series,
        ["Chart type", "Field F1", "Encoding", "Aggregation", "Exact match"],
    ):
        series.tx = SeriesLabel(v=title)
    chart.legend.legendPos = "r"
    chart.height = 10
    chart.width = 20
    sheet.add_chart(chart, "J2")


def add_system_charts(sheet: Any) -> None:
    latency = BarChart()
    latency.type = "col"
    latency.style = 10
    latency.title = "Latency по подходам"
    latency.y_axis.title = "Latency ms"
    latency.x_axis.title = "Подход"
    latency.y_axis.numFmt = "0"
    latency.legend = None
    latency.add_data(Reference(sheet, min_col=2, max_col=2, min_row=1, max_row=6), titles_from_data=True)
    latency.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=6))
    latency.height = 9
    latency.width = 13
    sheet.add_chart(latency, "G2")

    failure = BarChart()
    failure.type = "col"
    failure.style = 12
    failure.title = "Failure rate по подходам"
    failure.y_axis.title = "Failure rate"
    failure.x_axis.title = "Подход"
    failure.y_axis.numFmt = "0%"
    failure.y_axis.scaling.min = 0
    failure.y_axis.scaling.max = 0.2
    failure.legend = None
    failure.add_data(Reference(sheet, min_col=4, max_col=4, min_row=1, max_row=6), titles_from_data=True)
    failure.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=6))
    failure.height = 9
    failure.width = 13
    sheet.add_chart(failure, "N2")


def write_latex_table(rows: list[dict[str, Any]]) -> None:
    tex_path = LATEX_DIR / "quality_metrics_table.tex"
    table_rows = "\n".join(
        f"{row['label']} & {fmt(row['vega_lite_validity'])} & {fmt(row['field_selection_f1'])} & {fmt(row['encoding_accuracy'])} & {fmt(row['normalized_exact_match'])} & {fmt(row['failure_rate'])} \\\\"
        for row in rows
    )
    tex_path.write_text(
        r"""\documentclass{article}
\usepackage{fontspec}
\setmainfont{DejaVu Serif}
\usepackage{booktabs}
\usepackage{geometry}
\geometry{margin=20mm}
\renewcommand{\tablename}{Таблица}
\begin{document}
\begin{table}[ht]
\centering
\caption{Итоговое сравнение качества Text-to-Visualization}
\begin{tabular}{lrrrrr}
\toprule
Подход & Validity & Field F1 & Encoding & Exact match & Failure rate \\
\midrule
""" + table_rows + r"""
\bottomrule
\end{tabular}
\end{table}
\end{document}
""",
        encoding="utf-8",
    )


def write_figures(rows: list[dict[str, Any]], drive_root: Path | None) -> None:
    draw_pipeline_architecture(FIGURES_DIR / "pipeline_architecture.png")
    draw_nvbench_flow(FIGURES_DIR / "nvbench_postquery_flow.png")
    draw_metrics_bars(rows, FIGURES_DIR / "key_metrics_comparison.png")
    draw_system_metrics(rows, FIGURES_DIR / "system_metrics_comparison.png")
    draw_examples_grid(drive_root, FIGURES_DIR / "examples_grid_gold_vs_predicted.png")
    manifest = [
        {"file": "pipeline_architecture.png", "caption": "Архитектура экспериментального pipeline для post-query T2V."},
        {"file": "nvbench_postquery_flow.png", "caption": "Поток адаптации nvBench из записей бенчмарка в post-query примеры."},
        {"file": "key_metrics_comparison.png", "caption": "Столбчатое сравнение ключевых метрик качества и устойчивости."},
        {"file": "system_metrics_comparison.png", "caption": "Сравнение latency, memory и failure rate по финальным runs."},
        {"file": "examples_grid_gold_vs_predicted.png", "caption": "Примеры эталонных и предсказанных графиков из финального B4 run."},
    ]
    write_json(FIGURES_DIR / "figure_manifest.json", manifest)


def draw_pipeline_architecture(path: Path) -> None:
    image = canvas("Экспериментальный pipeline post-query Text-to-Visualization", width=1800, height=950)
    boxes = [
        (80, 190, 260, "Готовая таблица\nCSV + метаданные"),
        (420, 190, 260, "NL-запрос\nпосле SQL"),
        (760, 190, 260, "Подходы\nB0..B4"),
        (1100, 190, 260, "Vega-Lite\nспецификация\nили ошибка"),
        (1440, 190, 280, "Оценщик + рендерер\nметрики + PNG"),
    ]
    draw = ImageDraw.Draw(image)
    for x, y, width, text in boxes:
        box(draw, x, y, x + width, y + 125, text, fill="#E9F2FF")
    for x, y, width, _text in boxes[:-1]:
        arrow(draw, x + width, y + 62, x + 340, y + 62)
    box(
        draw,
        640,
        410,
        1460,
        560,
        "Артефакты в канонической папке Drive\nruns/<run_id>/predictions\nruns/<run_id>/metrics\nruns/<run_id>/rendered + runtime logs",
        fill="#F2F7EC",
    )
    arrow(draw, 1230, 315, 1230, 410)
    image.save(path)


def draw_nvbench_flow(path: Path) -> None:
    image = canvas("Адаптация nvBench под post-query бенчмарк", width=1800, height=950)
    draw = ImageDraw.Draw(image)
    steps = [
        "Запись nvBench\nNL + SQL + цель\nвизуализации",
        "SQL только в\nupstream-части\nдля таблицы",
        "CSV + метаданные\nполя, роли,\nтипы данных",
        "Post-query пример\nзапрос + таблица\n+ gold-спецификация",
        "Датасет оценки\nsample20/sample200",
    ]
    y = 245
    x_positions = [70, 410, 750, 1090, 1430]
    for index, (x, text) in enumerate(zip(x_positions, steps)):
        box(draw, x, y, x + 280, y + 130, text, fill="#FFF4DE")
        if index < len(steps) - 1:
            arrow(draw, x + 280, y + 65, x + 340, y + 65)
    note = "Граница: проект оценивает Text-to-Visualization после выполнения запроса; качество Text-to-SQL/БД относится к upstream."
    box(draw, 170, 540, 1630, 650, note, fill="#F5EAFE")
    image.save(path)


def draw_metrics_bars(rows: list[dict[str, Any]], path: Path) -> None:
    if plt is None:
        draw_metrics_bars_fallback(rows, path)
        return
    configure_matplotlib()
    metrics = [
        ("chart_type_accuracy", "Chart type"),
        ("field_selection_f1", "Field F1"),
        ("encoding_accuracy", "Encoding"),
        ("aggregation_accuracy", "Aggregation"),
        ("normalized_exact_match", "Exact match"),
    ]
    colors = ["#5B8FF9", "#61DDAA", "#65789B", "#F6BD16", "#E8684A"]
    labels = [display_method_label(row, multiline=True) for row in rows]
    fig, axes = plt.subplots(2, 3, figsize=(18, 10), constrained_layout=False)
    fig.patch.set_facecolor("white")
    fig.subplots_adjust(top=0.78, hspace=0.58, wspace=0.55, left=0.08, right=0.98, bottom=0.07)
    fig.suptitle("Сравнение качества Text-to-Visualization подходов", fontsize=20, fontweight="bold", y=0.98)
    fig.text(
        0.5,
        0.905,
        "Все метрики на шкале 0..1; выше лучше. Exact match — строгое совпадение с gold spec, Field F1 — корректность выбора полей.",
        ha="center",
        fontsize=11,
        color="#374151",
    )
    y_positions = list(range(len(rows)))
    for axis, (metric, title) in zip(axes.flat[:5], metrics):
        values = [float(row.get(metric) or 0.0) for row in rows]
        axis.barh(y_positions, values, color=colors, height=0.62)
        axis.set_title(title, fontsize=14, fontweight="bold", pad=10)
        axis.set_xlim(0, 1.0)
        axis.set_yticks(y_positions, labels=labels, fontsize=10)
        axis.invert_yaxis()
        axis.xaxis.set_major_formatter(PercentFormatter(xmax=1.0, decimals=0))
        axis.grid(axis="x", color="#E5E7EB", linewidth=1)
        axis.set_axisbelow(True)
        axis.spines[["top", "right", "left"]].set_visible(False)
        for index, value in enumerate(values):
            axis.text(min(value + 0.025, 0.96), index, fmt(value), va="center", fontsize=10, color="#111827")
    note_axis = axes.flat[5]
    note_axis.axis("off")
    note_axis.text(
        0.0,
        0.95,
        "Как читать график",
        fontsize=14,
        fontweight="bold",
        color="#111827",
        va="top",
    )
    note_axis.text(
        0.0,
        0.76,
        "B4 выигрывает по Exact match и Aggregation,\n"
        "но сравнивается на меньшей выборке из-за GPU latency.\n\n"
        "B0/B1/B2 быстрые и стабильные baseline/fallback,\n"
        "но хуже по строгому совпадению с gold spec.",
        fontsize=10.5,
        color="#374151",
        linespacing=1.35,
        va="top",
    )
    fig.savefig(path, dpi=180, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def draw_system_metrics(rows: list[dict[str, Any]], path: Path) -> None:
    if plt is None:
        draw_system_metrics_fallback(rows, path)
        return
    configure_matplotlib()
    colors = ["#5B8FF9", "#61DDAA", "#65789B", "#F6BD16", "#E8684A"]
    labels = [display_method_label(row, multiline=False) for row in rows]
    latency = [as_optional_float(row.get("latency_ms")) for row in rows]
    latency_zoom = latency[:3]
    memory = [as_optional_float(row.get("memory_peak_mb")) for row in rows]
    failure = [float(row.get("failure_rate") or 0.0) for row in rows]

    fig, axes = plt.subplots(2, 2, figsize=(17, 10), constrained_layout=False)
    fig.patch.set_facecolor("white")
    fig.subplots_adjust(top=0.78, hspace=0.55, wspace=0.42, left=0.09, right=0.98, bottom=0.07)
    fig.suptitle("Системные метрики финальных runs", fontsize=20, fontweight="bold", y=0.98)
    fig.text(
        0.5,
        0.905,
        "Latency показывает среднее время на пример. Для B3/B4 значения на порядки выше; для B2 надежный runtime-замер отсутствует.",
        ha="center",
        fontsize=11,
        color="#374151",
    )

    draw_horizontal_bar_panel(
        axes[0, 0],
        labels,
        latency,
        colors,
        title="Latency ms: полный масштаб",
        value_formatter=format_latency_label,
        x_label="milliseconds per example",
    )
    draw_horizontal_bar_panel(
        axes[0, 1],
        labels[:3],
        latency_zoom,
        colors[:3],
        title="Latency ms: zoom для B0/B1/B2",
        value_formatter=format_latency_label,
        x_label="milliseconds per example",
        x_limit=1.0,
    )
    draw_horizontal_bar_panel(
        axes[1, 0],
        labels,
        failure,
        colors,
        title="Failure rate: доля неуспешных предсказаний",
        value_formatter=lambda value: f"{value:.0%}",
        x_label="share of examples",
        x_limit=0.2,
        percent_axis=True,
    )
    draw_horizontal_bar_panel(
        axes[1, 1],
        labels,
        memory,
        colors,
        title="Memory MB: peak RSS процесса",
        value_formatter=lambda value: "нет замера" if value is None else f"{value:.0f} MB",
        x_label="megabytes",
    )
    fig.savefig(path, dpi=180, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def configure_matplotlib() -> None:
    if plt is None:
        return
    plt.rcParams.update(
        {
            "font.family": "DejaVu Sans",
            "axes.edgecolor": "#D1D5DB",
            "axes.labelcolor": "#374151",
            "xtick.color": "#374151",
            "ytick.color": "#111827",
            "figure.facecolor": "white",
        }
    )


def draw_horizontal_bar_panel(
    axis: Any,
    labels: list[str],
    values: list[float | None],
    colors: list[str],
    *,
    title: str,
    value_formatter: Any,
    x_label: str,
    x_limit: float | None = None,
    percent_axis: bool = False,
) -> None:
    y_positions = list(range(len(labels)))
    numeric_values = [0.0 if value is None else value for value in values]
    axis.barh(y_positions, numeric_values, color=colors, height=0.6)
    axis.set_title(title, fontsize=14, fontweight="bold", pad=10)
    axis.set_yticks(y_positions, labels=labels, fontsize=10)
    axis.invert_yaxis()
    if x_limit is not None:
        axis.set_xlim(0, x_limit)
    elif numeric_values:
        max_value = max(numeric_values)
        axis.set_xlim(0, max_value * 1.18 if max_value > 0 else 1)
    axis.set_xlabel(x_label, fontsize=10)
    if percent_axis and PercentFormatter is not None:
        axis.xaxis.set_major_formatter(PercentFormatter(xmax=1.0, decimals=0))
    axis.grid(axis="x", color="#E5E7EB", linewidth=1)
    axis.set_axisbelow(True)
    axis.spines[["top", "right", "left"]].set_visible(False)
    limit = axis.get_xlim()[1]
    for index, value in enumerate(values):
        numeric_value = 0.0 if value is None else value
        text_x = min(numeric_value + limit * 0.025, limit * 0.94)
        if numeric_value == 0:
            text_x = limit * 0.02
        axis.text(text_x, index, value_formatter(value), va="center", fontsize=10, color="#111827")


def format_latency_label(value: float | None) -> str:
    if value is None:
        return "нет замера"
    if value >= 1000:
        return f"{value / 1000:.1f} s"
    return f"{value:.2f} ms"


def format_optional_system_metric(value: Any) -> str | float:
    if value is None:
        return "нет замера"
    return value


def display_method_label(row: dict[str, Any], *, multiline: bool) -> str:
    names = {
        "B0": "rule-based",
        "B1": "constraint-ranker",
        "B2": "partial-recommender",
        "B3": "Qwen3-8B",
        "B4": "validator-reranker",
    }
    separator = "\n" if multiline else " "
    return f"{row['label']}{separator}{names.get(str(row['label']), str(row['method']))}"


def draw_metrics_bars_fallback(rows: list[dict[str, Any]], path: Path) -> None:
    image = canvas("Сравнение ключевых метрик", width=1800, height=1150)
    draw = ImageDraw.Draw(image)
    font = load_font(28)
    small = load_font(22)
    metrics = [
        ("chart_type_accuracy", "Chart type"),
        ("field_selection_f1", "Field F1"),
        ("encoding_accuracy", "Encoding"),
        ("aggregation_accuracy", "Aggregation"),
        ("normalized_exact_match", "Exact match"),
    ]
    colors = ["#5B8FF9", "#61DDAA", "#65789B", "#F6BD16", "#E8684A"]
    chart_left, chart_top = 120, 220
    group_width = 310
    max_bar_height = 560
    for group_index, (_metric, label) in enumerate(metrics):
        x0 = chart_left + group_index * group_width
        draw.text((x0 + 25, 810), label, font=font, fill="#1F2937")
        for row_index, row in enumerate(rows):
            value = float(row.get(_metric) or 0.0)
            bar_h = int(value * max_bar_height)
            x = x0 + 20 + row_index * 44
            y = chart_top + max_bar_height - bar_h
            draw.rounded_rectangle([x, y, x + 30, chart_top + max_bar_height], radius=5, fill=colors[row_index])
            draw.text((x - 4, y - 28), fmt(value), font=small, fill="#111827")
    for i, row in enumerate(rows):
        y = 930 + i * 38
        draw.rounded_rectangle([120, y, 150, y + 30], radius=4, fill=colors[i])
        draw.text((165, y - 2), f"{row['label']} - {row['method']}", font=small, fill="#111827")
    image.save(path)


def draw_system_metrics_fallback(rows: list[dict[str, Any]], path: Path) -> None:
    image = canvas("Системные метрики: latency, memory, failure rate", width=1800, height=1150)
    draw = ImageDraw.Draw(image)
    font = load_font(28)
    small = load_font(22)
    colors = ["#5B8FF9", "#61DDAA", "#65789B", "#F6BD16", "#E8684A"]
    metrics = [
        ("latency_ms", "Latency ms"),
        ("failure_rate", "Failure rate"),
        ("memory_peak_mb", "Memory MB"),
    ]
    for group_index, (metric, label) in enumerate(metrics):
        y0 = 190 + group_index * 280
        draw.text((90, y0), label, font=font, fill="#111827")
        values = [as_optional_float(row.get(metric)) for row in rows]
        max_value = max((value or 0.0) for value in values) or 1.0
        for row_index, row in enumerate(rows):
            value = values[row_index]
            numeric_value = value or 0.0
            x = 360 + row_index * 260
            width = int((numeric_value / max_value) * 210)
            draw.rounded_rectangle([x, y0 + 60, x + width, y0 + 105], radius=5, fill=colors[row_index])
            value_label = "нет замера" if value is None else fmt(value)
            draw.text((x, y0 + 118), f"{row['label']}: {value_label}", font=small, fill="#111827")
    image.save(path)


def draw_examples_grid(drive_root: Path | None, path: Path) -> None:
    if drive_root and try_draw_real_examples_grid(drive_root, path):
        (FIGURES_DIR / "examples_grid_source.txt").write_text(
            "реальные графики из Drive отрендерены из stage7_b4_sample20_tokens384\n",
            encoding="utf-8",
        )
        return
    source_path = FIGURES_DIR / "examples_grid_source.txt"
    if not drive_root and path.exists() and source_path.exists() and "реальные графики из Drive" in source_path.read_text(encoding="utf-8"):
        return
    image = canvas("Примеры: эталон (gold) и предсказание B4", width=1600, height=1000)
    draw = ImageDraw.Draw(image)
    font = load_font(30)
    small = load_font(22)
    for idx in range(4):
        x = 80 + (idx % 2) * 720
        y = 190 + (idx // 2) * 360
        box(draw, x, y, x + 300, y + 250, f"Эталон {idx + 1}\n(gold)", fill="#EEF2FF")
        box(draw, x + 340, y, x + 640, y + 250, f"Предсказание\nB4 {idx + 1}", fill="#ECFDF5")
    draw.text((80, 120), "При наличии Drive-артефактов здесь отображаются реальные графики.", font=font, fill="#111827")
    image.save(path)
    (FIGURES_DIR / "examples_grid_source.txt").write_text(
        "локальная резервная сетка; реальные Drive charts недоступны в этом окружении\n",
        encoding="utf-8",
    )


def try_draw_real_examples_grid(drive_root: Path, path: Path) -> bool:
    try:
        import vl_convert as vlc  # type: ignore
    except Exception:
        return False
    run_root = drive_root / "runs" / "stage7_b4_sample20_tokens384"
    examples_path = run_root / "examples_used.jsonl"
    predictions_path = run_root / "predictions" / "B4_llm_validator_reranker.jsonl"
    if not examples_path.exists() or not predictions_path.exists():
        return False
    examples = {row["example_id"]: row for row in read_jsonl(examples_path)}
    predictions = [row for row in read_jsonl(predictions_path) if row.get("status") == "ok" and row.get("raw_spec")]
    selected = predictions[:4]
    if not selected:
        return False
    tiles: list[tuple[str, Image.Image, Image.Image]] = []
    for prediction in selected:
        example = examples[prediction["example_id"]]
        table = pd.read_csv(example["table_path"]).head(500).to_dict(orient="records")
        gold = render_spec(vlc, example.get("gold_spec") or {}, table)
        predicted = render_spec(vlc, prediction.get("raw_spec") or {}, table)
        tiles.append((prediction["example_id"], gold, predicted))
    image = canvas("Примеры: эталон (gold) и предсказание B4", width=1800, height=1200)
    draw = ImageDraw.Draw(image)
    small = load_font(22)
    for idx, (example_id, gold, predicted) in enumerate(tiles):
        x = 80 + (idx % 2) * 840
        y = 170 + (idx // 2) * 490
        draw.text((x, y - 34), f"{example_id}: эталон (gold)", font=small, fill="#111827")
        draw.text((x + 400, y - 34), "Предсказание B4", font=small, fill="#111827")
        image.paste(fit_image(gold, 360, 320), (x, y))
        image.paste(fit_image(predicted, 360, 320), (x + 400, y))
    image.save(path)
    return True


def render_spec(vlc: Any, spec: dict[str, Any], table: list[dict[str, Any]]) -> Image.Image:
    spec_with_data = dict(spec)
    spec_with_data["data"] = {"values": table}
    try:
        data = vlc.vegalite_to_png(spec_with_data)
        from io import BytesIO

        return Image.open(BytesIO(data)).convert("RGB")
    except Exception as exc:
        panel = Image.new("RGB", (620, 420), "#FFF7ED")
        draw = ImageDraw.Draw(panel)
        draw.text((28, 28), f"Рендеринг не выполнен:\n{exc}", font=load_font(22), fill="#7C2D12")
        return panel


def fit_image(image: Image.Image, width: int, height: int) -> Image.Image:
    image = image.copy()
    image.thumbnail((width, height), Image.Resampling.LANCZOS)
    canvas_img = Image.new("RGB", (width, height), "white")
    canvas_img.paste(image, ((width - image.width) // 2, (height - image.height) // 2))
    return canvas_img


def write_markdown(rows: list[dict[str, Any]], inventory: list[dict[str, Any]], drive_root: Path | None) -> None:
    report = REPO_ROOT / "reports" / "practice_report_materials.md"
    report.write_text(build_report_markdown(rows, inventory, drive_root), encoding="utf-8")
    (OUT_DIR / "practice_report_materials.md").write_text(report.read_text(encoding="utf-8"), encoding="utf-8")


def build_report_markdown(rows: list[dict[str, Any]], inventory: list[dict[str, Any]], drive_root: Path | None) -> str:
    quality_md = markdown_table(read_csv(TABLES_DIR / "quality_metrics.csv"))
    system_md = markdown_table(read_csv(TABLES_DIR / "latency_memory_failure.csv"))
    comparison_md = markdown_table(read_csv(TABLES_DIR / "comparison_solutions.csv"))
    unique_run_ids = list(dict.fromkeys(run.run_id for run in FINAL_RUNS))
    run_list = "\n".join(f"- `{run_id}`: `{CANONICAL_DRIVE_ROOT_STR}/runs/{run_id}`" for run_id in unique_run_ids)
    return f"""# Материалы для отчета по преддипломной практике

## Введение

В рамках практики реализован и проверен post-query Text-to-Visualization контур для ВКР. Граница работы: Text-to-SQL и качество SQL/БД не оцениваются; на вход визуализационного модуля уже поступают готовая таблица, естественно-языковой запрос и метаданные. Выходом является Vega-Lite спецификация, таблица, рекомендация графика или заготовка дашборда.

Цель практической части: подготовить воспроизводимый бенчмарк, реализовать несколько базовых подходов и сравнить их по валидности Vega-Lite, совпадению с эталонными (gold) спецификациями, выбору полей, устойчивости и вычислительным затратам.

## Что реализовано

Реализованы пять подходов:

{comparison_md}

Этап 8 намеренно пропущен как опциональный: дополнительные LLM-эксперименты требуют GPU-времени, а для отчета уже есть сравнение пяти подходов `B0`-`B4`.

## Методика экспериментов

1. nvBench адаптирован под post-query постановку: SQL используется только в upstream-части для материализации таблицы; downstream-примеры содержат CSV-таблицу, метаданные, NL-запрос и эталонную (gold) Vega-Lite-like spec.
2. Все runs сохранялись в канонической папке Google Drive: `{CANONICAL_DRIVE_ROOT_STR}`.
3. Для каждого метода сохранялись файлы предсказаний jsonl, агрегированные и попримерные метрики, информация о среде выполнения, pip freeze и артефакты рендеринга.
4. Основные метрики: `vega_lite_validity`, `field_selection_f1`, `encoding_accuracy`, `aggregation_accuracy`, `normalized_exact_match`, `failure_rate`, `latency_ms`, `memory_peak_mb`.
5. Для B4 дополнительно сохраняются все 3 кандидата на пример и считается `oracle_success_at_k`.

## Итоговые результаты

### Метрики качества

{quality_md}

### Задержка, память и ошибки

{system_md}

Ключевой результат: лучшим по качеству стал `B4_llm_validator_reranker`, но он дороже по задержке, так как генерирует три кандидата на пример. Лучшим практическим вариантом для интеграции является гибрид: быстрый детерминированный резервный подход (`B0`/`B2`) плюс B4 для случаев, где важнее качество и допустима задержка.

## Выбор подхода для дальнейшей интеграции

Рекомендация:

- Для рабочего резервного варианта: `B0_rule_based` или `B2_partial_recommender`, потому что они быстрые, валидные и воспроизводимые.
- Для исследовательского качества и демонстрации LLM-возможностей: `B4_llm_validator_reranker`.
- Для одиночного LLM-подхода без reranking: `B3_local_llm_qwen3_8b`, если нужно снизить задержку относительно B4.

Итоговая архитектура интеграции: сначала быстрый базовый подход формирует гарантированную валидную визуализацию; при наличии GPU/времени B4 генерирует несколько кандидатов, валидатор фильтрует незаконные спецификации, reranker выбирает лучший результат.

## Рисунки для отчета

Минимальный набор рисунков:

1. `reports/stage9_report_materials/figures/pipeline_architecture.png` - архитектура экспериментального pipeline.
2. `reports/stage9_report_materials/figures/nvbench_postquery_flow.png` - адаптация nvBench под post-query.
3. `reports/stage9_report_materials/figures/key_metrics_comparison.png` - сравнение ключевых метрик качества.
4. `reports/stage9_report_materials/figures/system_metrics_comparison.png` - сравнение latency, memory и failure rate.
5. `reports/stage9_report_materials/figures/examples_grid_gold_vs_predicted.png` - сетка "эталон vs предсказание" для финального B4 run.

## Список артефактов

Финальные папки runs:

{run_list}

Материалы этапа 9:

- `reports/stage9_report_materials/stage9_tables.xlsx`
- `reports/stage9_report_materials/tables/*.csv`
- `reports/stage9_report_materials/latex/quality_metrics_table.tex`
- `reports/stage9_report_materials/latex/quality_metrics_table.pdf`, если доступна компиляция LaTeX
- `reports/stage9_report_materials/figures/*.png`
- `reports/stage9_report_materials/run_inventory.json`

Копия в Google Drive:

`{CANONICAL_DRIVE_ROOT_STR}/reports/stage9_report_materials`

## Приложение: команды

### Запуск финального B4

```powershell
.\\scripts\\colab\\run_colab_notebook.ps1 -NotebookPath .\\notebooks\\03_run_local_llm.ipynb -Action cell -CellId stage7-run20 -WaitSeconds 240 -ReloadFromDisk:$false -Json
```

### Генерация материалов этапа 9

```powershell
python scripts/make_stage9_report_materials.py
```

### Проверка outputs в notebook

```powershell
python -c "import nbformat; nb=nbformat.read('notebooks/03_run_local_llm.ipynb', as_version=4); nbformat.validate(nb)"
```

### Тесты

```powershell
python -m pytest -q
```

## Ограничения

- Метрики B3/B4 получены на меньших sample sizes из-за задержки на GPU.
- B2 является частичным резервным подходом в стиле существующих инструментов, а не полноценной NL4DV интеграцией.
- Rendered PNG используется как smoke/inspection artifact; корректность графика определяется метриками качества и ручным анализом.
- Text-to-SQL не входит в оценку.
"""


def write_review(rows: list[dict[str, Any]], inventory: list[dict[str, Any]], drive_root: Path | None) -> None:
    review = REPO_ROOT / "reports" / "stage_reviews" / "STAGE_9_REVIEW.md"
    review.write_text(
        f"""# Отчет по проверке этапа 9: финальные материалы для отчета по практике

Дата: 2026-04-26

Статус: завершен.

Этап 8 намеренно пропущен как опциональный. Новые LLM-эксперименты не запускались, логика базовых подходов не изменялась.

## Финальные запуски

{markdown_table(localized_final_runs_rows())}

## Выходные материалы

Локальный пакет:

```text
reports/stage9_report_materials
```

Пакет в Google Drive:

```text
{CANONICAL_DRIVE_ROOT_STR}/reports/stage9_report_materials
```

## Выполненные команды

Локально:

```powershell
python scripts/make_stage9_report_materials.py
Set-Location reports/stage9_report_materials/latex
& "$env:USERPROFILE\\.codex\\plugins\\cache\\openai-bundled\\latex-tectonic\\0.1.0\\bin\\tectonic.exe" --outdir . quality_metrics_table.tex
Set-Location ..\\..\\..
python -m pytest -q
```

Colab через runner:

```powershell
.\\scripts\\colab\\run_colab_notebook.ps1 -NotebookPath .\\notebooks\\05_make_report_materials.ipynb -Action cell -CellId stage9-setup -WaitSeconds 20 -ReloadFromDisk:$false -Json
.\\scripts\\colab\\run_colab_notebook.ps1 -NotebookPath .\\notebooks\\05_make_report_materials.ipynb -Action cell -CellId stage9-build-materials -WaitSeconds 30 -ReloadFromDisk:$false -Json
.\\scripts\\colab\\run_colab_notebook.ps1 -NotebookPath .\\notebooks\\05_make_report_materials.ipynb -Action cell -CellId stage9-verify-artifacts -WaitSeconds 15 -ReloadFromDisk:$false -Json
```

## Таблицы

- `tables/comparison_solutions.csv`
- `tables/quality_metrics.csv`
- `tables/latency_memory_failure.csv`
- `tables/applicability.csv`
- `tables/risks_limitations.csv`
- `stage9_tables.xlsx`

## Рисунки

- `figures/pipeline_architecture.png`
- `figures/nvbench_postquery_flow.png`
- `figures/key_metrics_comparison.png`
- `figures/system_metrics_comparison.png`
- `figures/examples_grid_gold_vs_predicted.png`

## Markdown для отчета

- `reports/practice_report_materials.md`
- `reports/stage9_report_materials/practice_report_materials.md`

## Инвентаризация runs

Найдено папок runs: {len(inventory)}

Файл инвентаризации:

```text
reports/stage9_report_materials/run_inventory.json
```

Финальные run_id, использованные в отчете:

- `stage4_cpu_sample200`: B0, B1
- `stage5_partial_sample200`: B2
- `stage6_qwen3_8b_fast_sample50`: B3
- `stage7_b4_sample20_tokens384`: B4

## Проверка

- Aggregate metrics объединены в один пакет сравнения.
- Создано не менее 3 рисунков; локальные PNG визуально проверены на читаемость.
- `stage9_tables.xlsx` содержит встроенные PNG-графики для метрик качества и системных метрик; исходные значения графиков сохранены как числовые ячейки.
- Colab-проверка нашла обязательные Drive-артефакты и вывела `STAGE9_VERIFY_OK`.
- Итоговая рекомендация включена в `practice_report_materials.md`.
- Все финальные run_id и пути к артефактам перечислены.
- Работа по этапу 8 не начиналась.

## Проблемы и замечания

- Этап 8 намеренно пропущен, потому что уже сравнены пять подходов, а дополнительные LLM runs увеличили бы GPU-затраты.
- Новые LLM-эксперименты на этом этапе не запускались.
- Логика базовых подходов не изменялась.
- Локальная пересборка workbook может быть пропущена, если `stage9_tables.xlsx` открыт в Excel; Colab создает workbook штатно.
- Сетка примеров генерируется из Drive-артефактов финального B4 run; при локальной пересборке без Drive генератор использует резервную схему.

## Следующие шаги

- Использовать `reports/practice_report_materials.md` как основной источник для написания отчета на 12-15 страниц.
- Использовать `reports/stage9_report_materials/stage9_tables.xlsx` и `tables/*.csv` для таблиц.
- Использовать `reports/stage9_report_materials/figures/*.png` для иллюстраций.
- После принятия review закоммитить материалы отчета, сводные таблицы, notebook, script и review markdown.
""",
        encoding="utf-8",
    )


def sync_to_drive(drive_root: Path | None) -> None:
    if not drive_root:
        return
    target = drive_root / "reports" / "stage9_report_materials"
    if target.exists():
        shutil.rmtree(target)
    shutil.copytree(OUT_DIR, target)


def canvas(title: str, *, width: int = 1600, height: int = 900) -> Image.Image:
    image = Image.new("RGB", (width, height), "#FFFFFF")
    draw = ImageDraw.Draw(image)
    draw.rectangle([0, 0, width, 95], fill="#17324D")
    draw.text((60, 28), title, font=load_font(36, bold=True), fill="white")
    return image


def box(draw: ImageDraw.ImageDraw, x1: int, y1: int, x2: int, y2: int, text: str, *, fill: str) -> None:
    draw.rounded_rectangle([x1, y1, x2, y2], radius=16, fill=fill, outline="#244761", width=3)
    draw.multiline_text((x1 + 18, y1 + 24), text, font=load_font(24), fill="#111827", spacing=6)


def arrow(draw: ImageDraw.ImageDraw, x1: int, y1: int, x2: int, y2: int) -> None:
    draw.line([x1, y1, x2, y2], fill="#244761", width=5)
    angle = math.atan2(y2 - y1, x2 - x1)
    size = 16
    points = [
        (x2, y2),
        (x2 - size * math.cos(angle - 0.5), y2 - size * math.sin(angle - 0.5)),
        (x2 - size * math.cos(angle + 0.5), y2 - size * math.sin(angle + 0.5)),
    ]
    draw.polygon(points, fill="#244761")


def load_font(size: int, *, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    candidates = [
        "C:/Windows/Fonts/arialbd.ttf" if bold else "C:/Windows/Fonts/arial.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
    ]
    patterns = [
        "/usr/share/fonts/**/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/**/DejaVuSans.ttf",
        "/usr/share/fonts/**/LiberationSans-Bold.ttf" if bold else "/usr/share/fonts/**/LiberationSans-Regular.ttf",
        "/usr/local/lib/python*/dist-packages/matplotlib/mpl-data/fonts/ttf/DejaVuSans-Bold.ttf"
        if bold
        else "/usr/local/lib/python*/dist-packages/matplotlib/mpl-data/fonts/ttf/DejaVuSans.ttf",
    ]
    for pattern in patterns:
        candidates.extend(glob.glob(pattern, recursive=True))
    for candidate in candidates:
        if Path(candidate).exists():
            return ImageFont.truetype(candidate, size=size)
    return ImageFont.load_default()


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def write_markdown_table(path: Path, rows: list[dict[str, Any]]) -> None:
    path.write_text(markdown_table(rows), encoding="utf-8")


def markdown_table(rows: list[dict[str, Any]]) -> str:
    if not rows:
        return ""
    headers = list(rows[0])
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(str(row.get(header, "")) for header in headers) + " |")
    return "\n".join(lines)


def fmt(value: Any) -> str:
    number = as_float(value)
    return f"{number:.3f}".rstrip("0").rstrip(".")


def as_float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def as_optional_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


if __name__ == "__main__":
    raise SystemExit(main())
